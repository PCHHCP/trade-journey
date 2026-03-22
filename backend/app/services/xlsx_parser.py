import logging
import re
import warnings
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.exceptions import ValidationException

logger = logging.getLogger(__name__)

METADATA_LABEL_ALIASES = {
    "account_name": ("名称", "name"),
    "account_number": ("账户", "账号", "帐号", "account", "login"),
    "company": ("公司", "company"),
}

HEADER_ALIASES = {
    "open_time": ("开仓时间", "open time"),
    "close_time": ("平仓时间", "close time"),
    "time": ("时间", "time"),
    "ticket": ("持仓", "票号", "ticket", "position"),
    "symbol": ("交易品种", "品种", "symbol"),
    "type": ("类型", "trade type", "type"),
    "volume": ("交易量", "手数", "volume", "lots"),
    "open_price": ("开仓价", "open price"),
    "close_price": ("平仓价", "close price"),
    "price": ("价位", "price"),
    "stop_loss": ("止损", "s/l", "sl", "stop loss"),
    "take_profit": ("止盈", "t/p", "tp", "take profit"),
    "profit": ("盈利", "profit"),
    "commission": ("手续费", "commission"),
    "swap": ("库存费", "swap"),
}

SECTION_MARKERS = ("持仓", "positions", "deals")
STOP_SECTION_MARKERS = ("订单", "orders")


@dataclass
class ParsedReport:
    account_name: str
    account_number: str
    platform: str
    company: str | None
    trades: list[dict[str, object]] = field(default_factory=list)


def _parse_number(value: object) -> Decimal | None:
    """Parse a number that may contain spaces as thousands separators."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    # Remove spaces used as thousands separators
    text = text.replace(" ", "").replace("\u00a0", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_datetime(value: object) -> datetime | None:
    """Parse datetime from MT5 report format: '2026.03.06 12:48:59'."""
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=UTC)
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        dt = datetime.strptime(text, "%Y.%m.%d %H:%M:%S")
        return dt.replace(tzinfo=UTC)
    except ValueError:
        return None


def _cell_text(cell: Cell) -> str:
    """Get cell value as stripped text."""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _normalize_text(value: object) -> str:
    text = str(value).strip().replace("：", ":").replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_label(value: object) -> str:
    text = _normalize_text(value).lower()
    return text.replace(":", "").strip()


def _extract_inline_label_value(text: str, aliases: tuple[str, ...]) -> str | None:
    normalized = _normalize_text(text)
    if ":" not in normalized:
        return None

    prefix, suffix = normalized.split(":", 1)
    if _normalize_label(prefix) in {_normalize_label(alias) for alias in aliases}:
        cleaned = suffix.strip()
        return cleaned or None

    return None


def _iter_non_empty_row_texts(row: tuple[Cell, ...], start_index: int = 0) -> list[str]:
    values: list[str] = []
    for cell in row[start_index:]:
        text = _cell_text(cell)
        if text:
            values.append(text)
    return values


def _parse_account_value(text: str) -> tuple[str, str]:
    cleaned = _normalize_text(text)
    match = re.match(r"(\d+)\s*(?:\((.+)\))?$", cleaned)
    if match:
        return match.group(1), (match.group(2) or "").strip()
    return cleaned, ""


def _extract_metadata(ws: Worksheet) -> tuple[str, str, str, str | None]:
    """Extract account metadata from the top rows of the report.

    Returns (account_name, account_number, platform, company).
    """
    account_name = ""
    account_number = ""
    platform = ""
    company: str | None = None

    for row in ws.iter_rows(min_row=1, max_row=30, max_col=12):
        for index, cell in enumerate(row):
            text = _cell_text(cell)
            if not text:
                continue

            account_name_value = _extract_inline_label_value(
                text,
                METADATA_LABEL_ALIASES["account_name"],
            )
            if account_name_value:
                account_name = account_name_value
                continue

            account_value = _extract_inline_label_value(
                text,
                METADATA_LABEL_ALIASES["account_number"],
            )
            if account_value:
                account_number, parsed_platform = _parse_account_value(account_value)
                if parsed_platform:
                    platform = parsed_platform
                continue

            company_value = _extract_inline_label_value(text, METADATA_LABEL_ALIASES["company"])
            if company_value:
                company = company_value or None
                continue

            normalized_label = _normalize_label(text)
            if normalized_label in {
                _normalize_label(alias)
                for alias in METADATA_LABEL_ALIASES["account_name"]
            }:
                next_values = _iter_non_empty_row_texts(row, index + 1)
                if next_values:
                    account_name = _normalize_text(next_values[0])
                continue

            if normalized_label in {
                _normalize_label(alias)
                for alias in METADATA_LABEL_ALIASES["account_number"]
            }:
                next_values = _iter_non_empty_row_texts(row, index + 1)
                if next_values:
                    joined_value = " ".join(next_values[:2])
                    account_number, parsed_platform = _parse_account_value(joined_value)
                    if parsed_platform:
                        platform = parsed_platform
                continue

            if normalized_label in {
                _normalize_label(alias) for alias in METADATA_LABEL_ALIASES["company"]
            }:
                next_values = _iter_non_empty_row_texts(row, index + 1)
                if next_values:
                    company = _normalize_text(next_values[0]) or None
                continue

    if not account_name or not account_number:
        raise ValidationException("Cannot extract account metadata from XLSX report")

    return account_name, account_number, platform, company


def _find_positions_section(ws: Worksheet) -> int | None:
    """Find the row number of the positions header row."""
    for row in ws.iter_rows(min_row=1):
        header_keys = {
            mapped_key
            for cell in row
            if (mapped_key := _map_single_header(_cell_text(cell))) is not None
        }
        if len(header_keys) >= 4 and ("ticket" in header_keys or "time" in header_keys):
            return int(row[0].row)
    return None


def _map_single_header(text: str) -> str | None:
    normalized = _normalize_label(text)
    for key, aliases in HEADER_ALIASES.items():
        if normalized in {_normalize_label(alias) for alias in aliases}:
            return key
    if any(marker in normalized for marker in SECTION_MARKERS):
        return "section_marker"
    return None


def _map_header_columns(header_row: tuple[Cell, ...]) -> dict[str, int]:
    """Map header names to column indices."""
    mapping: dict[str, int] = {}
    price_count = 0
    time_count = 0

    for i, cell in enumerate(header_row):
        key = _map_single_header(_cell_text(cell))
        if key in ("open_time", "close_time"):
            mapping[key] = i
        elif key == "time":
            # First shared "time" is open_time, second is close_time
            if time_count == 0:
                mapping["open_time"] = i
            else:
                mapping["close_time"] = i
            time_count += 1
        elif key == "ticket":
            mapping["ticket"] = i
        elif key == "symbol":
            mapping["symbol"] = i
        elif key == "type":
            mapping["type"] = i
        elif key == "volume":
            mapping["volume"] = i
        elif key in ("open_price", "close_price"):
            mapping[key] = i
        elif key == "price":
            # First shared "price" is open_price, second is close_price
            if price_count == 0:
                mapping["open_price"] = i
            else:
                mapping["close_price"] = i
            price_count += 1
        elif key == "stop_loss":
            mapping["stop_loss"] = i
        elif key == "take_profit":
            mapping["take_profit"] = i
        elif key == "profit":
            mapping["profit"] = i
        elif key == "commission":
            mapping["commission"] = i
        elif key == "swap":
            mapping["swap"] = i

    return mapping


def _parse_trade_type(value: object) -> str:
    """Parse trade type from MT5 format."""
    text = str(value).strip().lower() if value else ""
    if "buy" in text:
        return "buy"
    if "sell" in text:
        return "sell"
    return text


def _is_stop_section(text: str) -> bool:
    normalized = _normalize_label(text)
    return any(marker in normalized for marker in STOP_SECTION_MARKERS)


def parse_trade_report(file_content: bytes) -> ParsedReport:
    """Parse an MT5 XLSX trade history report.

    Extracts account metadata and the 持仓 (positions/deals) section.
    """
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style, apply openpyxl's default",
                category=UserWarning,
            )
            wb = load_workbook(
                filename=BytesIO(file_content),
                read_only=True,
                data_only=True,
            )
    except Exception as e:
        raise ValidationException(f"Cannot read XLSX file: {e}")

    ws = wb.active
    if ws is None:
        raise ValidationException("XLSX file has no active worksheet")

    account_name, account_number, platform, company = _extract_metadata(ws)

    header_row_num = _find_positions_section(ws)
    if header_row_num is None:
        raise ValidationException("Cannot find 持仓 (positions) section in report")

    # Get header row to map columns
    header_cells: tuple[Cell, ...] | None = None
    for row in ws.iter_rows(min_row=header_row_num, max_row=header_row_num):
        header_cells = row
        break

    if header_cells is None:
        raise ValidationException("Cannot read header row")

    col_map = _map_header_columns(header_cells)

    required_cols = ["ticket", "symbol", "type", "volume", "open_price", "open_time"]
    missing = [c for c in required_cols if c not in col_map]
    if missing:
        raise ValidationException(f"Missing required columns: {', '.join(missing)}")

    trades: list[dict[str, object]] = []

    for row in ws.iter_rows(min_row=header_row_num + 1):
        cells = row

        # Stop at next section header ("订单") or empty rows
        first_text = _cell_text(cells[0]) if cells else ""
        if _is_stop_section(first_text):
            break

        # Check if the ticket column has a value (skip empty/summary rows)
        ticket_val = cells[col_map["ticket"]].value if "ticket" in col_map else None
        if ticket_val is None:
            continue

        try:
            ticket_num = int(ticket_val)
        except (ValueError, TypeError):
            continue

        def _get_num(
            key: str,
            default: Decimal | None = None,
        ) -> Decimal | None:
            if key not in col_map:
                return default
            return _parse_number(cells[col_map[key]].value) or default

        def _get_dt(key: str) -> datetime | None:
            if key not in col_map:
                return None
            return _parse_datetime(cells[col_map[key]].value)

        open_time = _get_dt("open_time")
        close_time = _get_dt("close_time")

        if open_time is None or close_time is None:
            continue

        trade_data: dict[str, object] = {
            "ticket": ticket_num,
            "symbol": str(cells[col_map["symbol"]].value or "").strip(),
            "type": _parse_trade_type(cells[col_map["type"]].value),
            "volume": _get_num("volume", Decimal("0")),
            "open_price": _get_num("open_price", Decimal("0")),
            "stop_loss": _get_num("stop_loss"),
            "take_profit": _get_num("take_profit"),
            "open_time": open_time,
            "close_time": close_time,
            "close_price": _get_num("close_price", Decimal("0")),
            "commission": _get_num("commission", Decimal("0")),
            "swap": _get_num("swap", Decimal("0")),
            "profit": _get_num("profit", Decimal("0")),
        }

        trades.append(trade_data)

    wb.close()

    logger.info(
        "Parsed XLSX report",
        extra={"account_number": account_number, "trade_count": len(trades)},
    )

    return ParsedReport(
        account_name=account_name,
        account_number=account_number,
        platform=platform,
        company=company,
        trades=trades,
    )
